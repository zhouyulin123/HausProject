# -*- coding: utf-8 -*-
"""从 Excel 导入自家产品到商品库。

用法：
    python import_products.py --template          # 生成空白模板 products_import.xlsx
    python import_products.py products_import.xlsx  # 导入（按 SKU 去重：已存在则更新，否则新增）

Excel 有两个工作表：
    「成品家具」：sku, 名称, 类别, 空间, 风格, 材质, 价格, 价格上限, 尺寸, 卖点, 替代选择
    「定制报价」：项目名, 分类, 计价单位, 材料档位, 单价, 说明
"""

import sys

from openpyxl import Workbook, load_workbook

from app.db.database import Base, SessionLocal, engine
from app.db.models import CustomQuoteRule, Product

PRODUCT_HEADERS = ["sku", "名称", "类别", "空间", "风格", "材质", "价格", "价格上限", "尺寸", "卖点", "替代选择"]
RULE_HEADERS = ["项目名", "分类", "计价单位", "材料档位", "单价", "说明"]

TEMPLATE_FILE = "products_import.xlsx"


def make_template() -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "成品家具"
    ws1.append(PRODUCT_HEADERS)
    ws1.append(["SF-100", "示例：三人位布艺沙发", "沙发", "客厅", "奶油风",
                "科技布", 4999, 6999, "宽 2.4m", "耐抓易清洁，宠物家庭首选", "棉麻款（低 500 元）"])
    ws2 = wb.create_sheet("定制报价")
    ws2.append(RULE_HEADERS)
    ws2.append(["定制衣柜", "柜类定制", "㎡", "E0 颗粒板", 680, "投影面积计价，含基础五金"])
    for ws in (ws1, ws2):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16
    wb.save(TEMPLATE_FILE)
    print(f"模板已生成: {TEMPLATE_FILE}（示例行导入时会一并写入，请替换/删除）")


def _cell(row, idx):
    v = row[idx] if idx < len(row) else None
    if isinstance(v, str):
        v = v.strip()
    return v if v not in ("", None) else None


def import_file(path: str) -> None:
    Base.metadata.create_all(bind=engine)
    wb = load_workbook(path, data_only=True)
    db = SessionLocal()
    added = updated = rules_added = 0
    try:
        if "成品家具" in wb.sheetnames:
            for row in wb["成品家具"].iter_rows(min_row=2, values_only=True):
                if not _cell(row, 1):  # 名称为空跳过
                    continue
                data = dict(
                    sku=_cell(row, 0), name=_cell(row, 1), category=_cell(row, 2),
                    room=_cell(row, 3), style=_cell(row, 4), material=_cell(row, 5),
                    price=int(_cell(row, 6) or 0), price_max=(int(_cell(row, 7)) if _cell(row, 7) else None),
                    size=_cell(row, 8), selling_point=_cell(row, 9), alternative=_cell(row, 10),
                )
                existing = (
                    db.query(Product).filter(Product.sku == data["sku"]).first()
                    if data["sku"] else None
                )
                if existing:
                    for k, v in data.items():
                        setattr(existing, k, v)
                    existing.is_active = True
                    updated += 1
                else:
                    db.add(Product(**data))
                    added += 1

        if "定制报价" in wb.sheetnames:
            for row in wb["定制报价"].iter_rows(min_row=2, values_only=True):
                if not _cell(row, 0):
                    continue
                db.add(CustomQuoteRule(
                    project_name=_cell(row, 0), category=_cell(row, 1),
                    pricing_unit=_cell(row, 2) or "㎡", material_grade=_cell(row, 3),
                    unit_price=int(_cell(row, 4) or 0), description=_cell(row, 5),
                ))
                rules_added += 1

        db.commit()
        print(f"OK: 成品新增 {added}、更新 {updated}；定制规则新增 {rules_added}")
    finally:
        db.close()


if __name__ == "__main__":
    if "--template" in sys.argv:
        make_template()
    elif len(sys.argv) > 1:
        import_file(sys.argv[1])
    else:
        print(__doc__)
